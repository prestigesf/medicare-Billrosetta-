"""Load CMS data files into the engine's types.

CMS publishes the RVU, GPCI, and ZIP-locality files in layouts that vary by
year and by release format (CSV, fixed-width text, spreadsheet export). Rather
than hardcode one guessed layout, the mapping from file columns to fields is
*configuration*: a ColumnMap. Adopting a new file, or a new year's format, is a
ColumnMap change, not a code change.

That keeps invariant 2 intact — no CMS-specific layout knowledge lives in
executable engine code — and it means the first real file costs a config, not
a rewrite.

Every loader validates as it reads. A row that cannot be parsed is reported
with its line number and rejected. Nothing is silently coerced or defaulted,
because a bad rate that loads quietly is worse than a file that refuses.
"""
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Sequence, Tuple, Union

from .models import GPCI, RVUs

# How a field is addressed in a file:
#   "HCPCS"                     header name
#   5                           column position, 0-based
#   [start, end]                fixed-width character slice
#   {"join": [0, 2], "sep": "-"}  several columns combined into one key
#
# Position matters because CMS's RVU file splits its header across four rows
# and repeats names within one of them — "PE RVU" appears twice, once for
# non-facility and once for facility. Only position tells them apart.
FieldSpec = Union[str, int, Tuple[int, int], Dict[str, object]]

# How many bad rows to list before truncating. A malformed file can produce
# thousands; the first handful is what tells you what went wrong.
MAX_PROBLEMS_SHOWN = 20


class FileFormatError(Exception):
    """The file could not be read under the supplied ColumnMap."""

    def __init__(self, path: Path, problems: Sequence[str]):
        self.path = path
        self.problems = list(problems)
        shown = "\n  ".join(self.problems[:MAX_PROBLEMS_SHOWN])
        hidden = len(self.problems) - MAX_PROBLEMS_SHOWN
        more = f"\n  ... and {hidden} more" if hidden > 0 else ""
        super().__init__(f"{path.name}: {len(self.problems)} problem(s)\n  {shown}{more}")


@dataclass(frozen=True)
class ColumnMap:
    """Where each field lives in a particular file.

    fields: field name -> header name (delimited/xlsx) or (start, end) slice
    fixed_width: read by character offsets rather than delimiter
    skip_rows: leading rows to discard — CMS files carry several banner and
        copyright rows before the real header
    sheet: worksheet name for xlsx; None takes the active sheet
    """

    fields: Dict[str, FieldSpec]
    fixed_width: bool = False
    skip_rows: int = 0
    encoding: str = "utf-8-sig"
    sheet: Optional[str] = None

    def require(self, *names: str) -> None:
        missing = [n for n in names if n not in self.fields]
        if missing:
            raise ValueError(f"ColumnMap is missing required field(s): {missing}")

    @classmethod
    def from_json(cls, path: Union[str, Path]) -> "ColumnMap":
        """Load a layout from a JSON file.

        Layouts are data, so adopting a new CMS release means editing JSON, not
        Python. Fixed-width offsets are written as two-element lists.
        """
        spec = json.loads(Path(path).read_text())
        fields = {
            name: tuple(value) if isinstance(value, list) else value
            for name, value in spec["fields"].items()
        }
        return cls(
            fields=fields,
            fixed_width=spec.get("fixed_width", False),
            skip_rows=spec.get("skip_rows", 0),
            encoding=spec.get("encoding", "utf-8-sig"),
            sheet=spec.get("sheet"),
        )


DEFAULT_JOIN_SEPARATOR = "-"


def _resolve_indices(path: Path, headers: Sequence[str], colmap: ColumnMap) -> Dict[str, dict]:
    """Resolve each mapped field to column positions.

    Returns {field: {"indices": [...], "sep": str}}. Names are matched
    case-insensitively with collapsed whitespace, since CMS varies both
    between releases. Positions are used as given.
    """
    def key(text: str) -> str:
        return " ".join(str(text).split()).upper()

    by_name: Dict[str, int] = {}
    for position, header in enumerate(headers):
        if header is None or not str(header).strip():
            continue
        by_name.setdefault(key(header), position)

    def one(spec, problems) -> Optional[int]:
        if isinstance(spec, bool):
            problems.append(f"{spec!r} is not a valid column reference")
            return None
        if isinstance(spec, int):
            if 0 <= spec < len(headers):
                return spec
            problems.append(f"column position {spec} is outside the file's {len(headers)} columns")
            return None
        position = by_name.get(key(spec))
        if position is None:
            problems.append(
                f"column {spec!r} not in file; available: {sorted(by_name)}"
            )
        return position

    resolved: Dict[str, dict] = {}
    problems: List[str] = []

    for name, spec in colmap.fields.items():
        if isinstance(spec, dict):
            parts = [one(s, problems) for s in spec.get("join", [])]
            if any(p is None for p in parts):
                continue
            resolved[name] = {
                "indices": parts,
                "sep": str(spec.get("sep", DEFAULT_JOIN_SEPARATOR)),
            }
            continue

        position = one(spec, problems)
        if position is not None:
            resolved[name] = {"indices": [position], "sep": ""}

    if problems:
        raise FileFormatError(path, problems)
    return resolved


def _assemble(row: Sequence, resolved: Dict[str, dict]) -> Dict[str, str]:
    """Pull one record out of a raw row using resolved positions."""
    def cell(index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        return "" if value is None else str(value).strip()

    return {
        name: spec["sep"].join(cell(i) for i in spec["indices"])
        for name, spec in resolved.items()
    }


def _xlsx_rows(path: Path, colmap: ColumnMap) -> Iterator[Tuple[int, Dict[str, str]]]:
    """Yield rows from a spreadsheet.

    read_only mode holds an open file handle, and a generator can be abandoned
    part-way — by an exception on a bad column, or by a caller that stops
    iterating. The workbook is closed in a finally block so the handle is
    released on every exit path, not just the happy one.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    rows = None
    try:
        sheet = book[colmap.sheet] if colmap.sheet else book.active

        rows = sheet.iter_rows(values_only=True)
        for _ in range(colmap.skip_rows):
            next(rows, None)

        header = next(rows, None)
        if header is None:
            raise FileFormatError(path, ["no header row found after skip_rows"])

        resolved = _resolve_indices(
            path, [str(h) if h is not None else "" for h in header], colmap
        )

        for offset, row in enumerate(rows, start=colmap.skip_rows + 2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            yield offset, _assemble(row, resolved)
    finally:
        # The row iterator holds its own handle into the zip archive. Closing
        # the workbook alone leaves that one open when iteration is abandoned
        # part-way, e.g. by a bad-column error after the header is read.
        if rows is not None:
            rows.close()
        book.close()


def _rows(path: Path, colmap: ColumnMap) -> Iterator[Tuple[int, Dict[str, str]]]:
    """Yield (line number, {field: raw text}) for each data row."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        yield from _xlsx_rows(path, colmap)
        return

    text = path.read_text(encoding=colmap.encoding)
    lines = text.splitlines()[colmap.skip_rows:]

    if colmap.fixed_width:
        for offset, line in enumerate(lines, start=colmap.skip_rows + 1):
            if not line.strip():
                continue
            yield offset, {
                name: line[spec[0]:spec[1]].strip()
                for name, spec in colmap.fields.items()
            }
        return

    rows = list(csv.reader(lines))
    if not rows:
        raise FileFormatError(path, ["file has no header row"])

    resolved = _resolve_indices(path, rows[0], colmap)

    for offset, row in enumerate(rows[1:], start=colmap.skip_rows + 2):
        if not any(str(cell).strip() for cell in row):
            continue
        yield offset, _assemble(row, resolved)


def _number(raw: str, *, allow_blank: bool = False) -> Optional[float]:
    """Parse a numeric cell. Blank means absent, not zero — when permitted."""
    cleaned = raw.replace(",", "").replace("$", "").strip()
    if not cleaned:
        if allow_blank:
            return None
        raise ValueError("blank")
    return float(cleaned)


def load_rvus(path: Union[str, Path], colmap: ColumnMap) -> Dict[str, RVUs]:
    """Read the RVU file (PPRRVU) into {cpt_code: RVUs}."""
    path = Path(path)
    colmap.require(
        "cpt_code", "work", "practice_expense_facility",
        "practice_expense_non_facility", "malpractice", "status_code",
    )

    table: Dict[str, RVUs] = {}
    problems: List[str] = []

    for line_no, raw in _rows(path, colmap):
        code = raw["cpt_code"].upper()
        if not code:
            continue
        # A code can appear several times with different modifiers — global,
        # professional component (26), technical component (TC) — each with
        # its own RVUs. They are separate priceable lines, not duplicates.
        modifier = raw.get("modifier", "")
        try:
            entry = RVUs(
                cpt_code=code,
                modifier=modifier,
                work=_number(raw["work"]),
                # A code priced in only one setting leaves the other blank.
                # That must stay None, never 0.0 — see MissingPracticeExpenseRVU.
                practice_expense_facility=_number(
                    raw["practice_expense_facility"], allow_blank=True
                ),
                practice_expense_non_facility=_number(
                    raw["practice_expense_non_facility"], allow_blank=True
                ),
                malpractice=_number(raw["malpractice"]),
                status_code=raw["status_code"].upper(),
            )
        except ValueError as exc:
            problems.append(f"line {line_no}, CPT {code}: {exc}")
            continue

        if not entry.status_code:
            problems.append(f"line {line_no}, CPT {entry.key}: blank status code")
            continue
        if entry.key in table and table[entry.key] != entry:
            problems.append(f"line {line_no}, CPT {entry.key}: conflicting duplicate row")
            continue

        table[entry.key] = entry

    if problems:
        raise FileFormatError(path, problems)
    if not table:
        raise FileFormatError(path, ["no RVU rows parsed"])
    return table


def load_gpcis(path: Union[str, Path], colmap: ColumnMap) -> Dict[str, GPCI]:
    """Read the GPCI file into {locality_id: GPCI}."""
    path = Path(path)
    colmap.require(
        "locality_id", "locality_name", "work", "practice_expense", "malpractice"
    )

    table: Dict[str, GPCI] = {}
    problems: List[str] = []

    for line_no, raw in _rows(path, colmap):
        locality = raw["locality_id"].strip()
        if not locality:
            continue
        try:
            entry = GPCI(
                locality_id=locality,
                locality_name=raw["locality_name"] or locality,
                work=_number(raw["work"]),
                practice_expense=_number(raw["practice_expense"]),
                malpractice=_number(raw["malpractice"]),
            )
        except ValueError as exc:
            problems.append(f"line {line_no}, locality {locality}: {exc}")
            continue

        if locality in table and table[locality] != entry:
            problems.append(f"line {line_no}, locality {locality}: conflicting duplicate row")
            continue

        table[locality] = entry

    if problems:
        raise FileFormatError(path, problems)
    if not table:
        raise FileFormatError(path, ["no GPCI rows parsed"])
    return table


def load_zip_crosswalk(path: Union[str, Path], colmap: ColumnMap) -> Dict[str, str]:
    """Read the ZIP-to-locality crosswalk into {zip: locality_id}."""
    path = Path(path)
    colmap.require("zip_code", "locality_id")

    table: Dict[str, str] = {}
    problems: List[str] = []

    for line_no, raw in _rows(path, colmap):
        zip_code = raw["zip_code"].strip()[:5]
        locality = raw["locality_id"].strip()
        if not zip_code:
            continue
        if not locality:
            problems.append(f"line {line_no}, ZIP {zip_code}: blank locality")
            continue
        if not zip_code.isdigit() or len(zip_code) != 5:
            problems.append(f"line {line_no}: {zip_code!r} is not a 5-digit ZIP")
            continue
        if zip_code in table and table[zip_code] != locality:
            problems.append(
                f"line {line_no}, ZIP {zip_code}: maps to both "
                f"{table[zip_code]} and {locality}"
            )
            continue
        table[zip_code] = locality

    if problems:
        raise FileFormatError(path, problems)
    if not table:
        raise FileFormatError(path, ["no crosswalk rows parsed"])
    return table


def load_conversion_factor(path: Union[str, Path], colmap: ColumnMap) -> float:
    """Read the conversion factor from the RVU file.

    CMS repeats the factor on every data row. Reading it from the file rather
    than hardcoding it is the difference between a number that stays correct
    when CMS republishes and one that silently goes stale.

    Every row must agree. A file carrying two different factors is ambiguous
    and is refused rather than resolved by picking one.
    """
    path = Path(path)
    colmap.require("conversion_factor")

    seen: Dict[float, int] = {}
    problems: List[str] = []

    for line_no, raw in _rows(path, colmap):
        text = raw["conversion_factor"]
        if not text:
            continue
        try:
            value = _number(text)
        except ValueError:
            problems.append(f"line {line_no}: conversion factor {text!r} is not a number")
            continue
        seen[value] = seen.get(value, 0) + 1

    if problems:
        raise FileFormatError(path, problems)
    if not seen:
        raise FileFormatError(path, ["no conversion factor found in any row"])
    if len(seen) > 1:
        raise FileFormatError(
            path,
            [f"file carries {len(seen)} different conversion factors: "
             + ", ".join(f"{v} on {n} row(s)" for v, n in sorted(seen.items()))],
        )
    return next(iter(seen))
